"""
app.py
------
HomeFi Web — Flask entry point with Supabase + Multi-user login
Run: python app.py
Then open: http://127.0.0.1:5000
"""
import os
import json
import hashlib
import secrets
import numpy as np
import pandas as pd
from datetime import datetime
from flask import (Flask, render_template, jsonify, request,
                   redirect, url_for, session, flash)
from supabase import create_client, Client

app = Flask(__name__)
app.secret_key = secrets.token_hex(32)   # sessions ke liye

# ── Supabase ──────────────────────────────────────────────────────────────────
SUPABASE_URL = "https://mhdhhyynritbuhpurdii.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im1oZGhoeXlucml0YnVocHVyZGlpIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODExNjc0MTUsImV4cCI6MjA5Njc0MzQxNX0.L9ob_H7C9uFql9eIeT6-P9EIkfeRyw7bOmvk0Hm88_g"
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ── Auth helpers ──────────────────────────────────────────────────────────────
def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

def get_current_user():
    return session.get("username")

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not get_current_user():
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

# ── Data helpers ──────────────────────────────────────────────────────────────
def get_df():
    user = get_current_user()
    if not user:
        return pd.DataFrame(columns=["date","category","type","amount","description"])
    response = supabase.table("expenses").select("*").eq("user_id", user).execute()
    data = response.data
    if not data:
        return pd.DataFrame(columns=["date","category","type","amount","description"])
    df = pd.DataFrame(data)
    df["date"]   = pd.to_datetime(df["date"], errors="coerce")
    df["amount"] = pd.to_numeric(df["amount"], errors="coerce").fillna(0)
    df = df.dropna(subset=["date"])
    return df

def add_month_col(df, col="date"):
    df = df.copy()
    df[col] = pd.to_datetime(df[col], errors="coerce")
    df["month"] = df[col].dt.to_period("M")
    return df

def safe_json(obj):
    if isinstance(obj, (np.integer,)):  return int(obj)
    if isinstance(obj, (np.floating,)): return float(obj)
    if isinstance(obj, np.ndarray):     return obj.tolist()
    return str(obj)

# ── Auth routes ───────────────────────────────────────────────────────────────
@app.route("/login", methods=["GET","POST"])
def login():
    if get_current_user():
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        username = request.form.get("username","").strip().lower()
        password = request.form.get("password","")
        result = supabase.table("users").select("*").eq("username", username).execute()
        if result.data and result.data[0]["password_hash"] == hash_password(password):
            session["username"] = username
            session["name"]     = result.data[0]["name"]
            return redirect(url_for("dashboard"))
        flash("Galat username ya password!", "error")
    return render_template("login.html")

@app.route("/signup", methods=["GET","POST"])
def signup():
    if get_current_user():
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        name     = request.form.get("name","").strip()
        username = request.form.get("username","").strip().lower()
        password = request.form.get("password","")
        if not name or not username or not password:
            flash("Saare fields fill karo!")
            return render_template("signup.html")
        existing = supabase.table("users").select("username").eq("username", username).execute()
        if existing.data:
            flash("Ye username already le liya gaya hai!")
            return render_template("signup.html")
        supabase.table("users").insert({
            "username": username,
            "name": name,
            "password_hash": hash_password(password)
        }).execute()
        flash(f"Account ban gaya! Ab login karo.", "success")
        return redirect(url_for("login"))
    return render_template("signup.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ── Main routes ───────────────────────────────────────────────────────────────
@app.route("/")
@login_required
def dashboard():
    df = get_df()
    # Filter to current month only
    current_month = pd.Timestamp.now().to_period("M")
    df = add_month_col(df)
    df = df[df["month"] == current_month].drop(columns=["month"])
    if df.empty:
        stats = dict(total_income=0, total_expense=0, net_savings=0,
                     savings_rate=0, num_transactions=0, months=0,
                     monthly_labels=[], monthly_exp=[], monthly_inc=[])
        return render_template("dashboard.html", stats=stats,
                       recent=[], months="[]", incomes="[]",
                       expenses="[]", savings="[]",
                       cat_names="[]", cat_amounts="[]",
                       username=get_current_user(), name=session.get("name",""))

    exp_df = add_month_col(df[df["type"]=="expense"].copy())
    inc_df = add_month_col(df[df["type"]=="income"].copy())

    monthly_exp = exp_df.groupby("month")["amount"].sum()
    monthly_inc = inc_df.groupby("month")["amount"].sum()
    all_months  = sorted(set(monthly_exp.index) | set(monthly_inc.index))

    total_income  = float(df[df["type"]=="income"]["amount"].sum())
    total_expense = float(df[df["type"]=="expense"]["amount"].sum())
    net_savings   = total_income - total_expense
    savings_rate  = round((net_savings/total_income)*100, 1) if total_income else 0
    months_count  = pd.to_datetime(df["date"], errors="coerce").dt.to_period("M").nunique()

    stats = dict(
        total_income=total_income, total_expense=total_expense,
        net_savings=net_savings, savings_rate=savings_rate,
        num_transactions=len(df), months=int(months_count),
        monthly_labels=[str(m) for m in all_months],
        monthly_exp=[float(monthly_exp.get(m,0)) for m in all_months],
        monthly_inc=[float(monthly_inc.get(m,0)) for m in all_months],
    )
    # Recent transactions
    recent_raw = supabase.table("expenses").select("*").eq("user_id", get_current_user()).order("date", desc=True).limit(10).execute()
    recent = recent_raw.data or []

# Category data
    cat_data = df[df["type"]=="expense"].groupby("category")["amount"].sum()

    return render_template("dashboard.html", stats=stats,
                       recent=recent,
                       months=json.dumps([str(m) for m in all_months]),
                       incomes=json.dumps([float(monthly_inc.get(m,0)) for m in all_months]),
                       expenses=json.dumps([float(monthly_exp.get(m,0)) for m in all_months]),
                       savings=json.dumps([float(monthly_inc.get(m,0)-monthly_exp.get(m,0)) for m in all_months]),
                       cat_names=json.dumps(cat_data.index.tolist()),
                       cat_amounts=json.dumps([float(v) for v in cat_data.values]),
                       username=get_current_user(), name=session.get("name",""))

@app.route("/add", methods=["GET","POST"])
@login_required
def add_entry():
    categories = ["Food","Transport","Entertainment","Shopping",
                  "Healthcare","Utilities","Education","Rent","Salary","Other"]
    if request.method == "POST":
        supabase.table("expenses").insert({
            "date":        request.form["date"],
            "category":    request.form["category"],
            "type":        request.form["type"],
            "amount":      float(request.form["amount"]),
            "description": request.form.get("description",""),
            "user_id":     get_current_user()
        }).execute()
        return redirect(url_for("dashboard"))
    return render_template("add.html", categories=categories)

@app.route("/categories")
@login_required
def categories(): return render_template("categories.html")

@app.route("/trends")
@login_required
def trends(): return render_template("trends.html")

@app.route("/budget")
@login_required
def budget(): return render_template("budget.html")

@app.route("/forecast")
@login_required
def forecast(): return render_template("forecast.html")

@app.route("/recommendations")
@login_required
def recommendations(): return render_template("recommendations.html")

# ── API routes ────────────────────────────────────────────────────────────────
@app.route("/api/categories")
@login_required
def api_categories():
    df = get_df()
    if df.empty: return jsonify({"labels":[],"values":[]})
    exp = df[df["type"]=="expense"].groupby("category")["amount"].sum()
    return jsonify({"labels": exp.index.tolist(), "values": [float(v) for v in exp.values]})

@app.route("/api/trends")
@login_required
def api_trends():
    df = get_df()
    if df.empty: return jsonify({"months":[],"income":[],"expense":[],"savings":[]})
    exp_df = add_month_col(df[df["type"]=="expense"].copy())
    inc_df = add_month_col(df[df["type"]=="income"].copy())
    me = exp_df.groupby("month")["amount"].sum()
    mi = inc_df.groupby("month")["amount"].sum()
    all_m = sorted(set(me.index)|set(mi.index))
    return jsonify({
        "months":  [str(m) for m in all_m],
        "income":  [float(mi.get(m,0)) for m in all_m],
        "expense": [float(me.get(m,0)) for m in all_m],
        "savings": [float(mi.get(m,0)-me.get(m,0)) for m in all_m],
    })

@app.route("/api/budget")
@login_required
def api_budget():
    df = get_df()
    budgets_path = os.path.join("Data","budgets.json")
    if not os.path.exists(budgets_path):
        return jsonify({"error":"No budget set yet. Add a Data/budgets.json file."})
    with open(budgets_path) as f:
        budgets = json.load(f)
    if df.empty:
        return jsonify({"categories":list(budgets.keys()),
                        "spent":[0]*len(budgets),
                        "budget":list(budgets.values())})
    exp = df[df["type"]=="expense"].groupby("category")["amount"].sum()
    cats = list(budgets.keys())
    return jsonify({
        "categories": cats,
        "spent":  [float(exp.get(c,0)) for c in cats],
        "budget": [float(budgets[c]) for c in cats],
    })

@app.route("/api/forecast")
@login_required
def api_forecast():
    df = get_df()
    if df.empty: return jsonify({"months":[],"forecast":[],"actual":[]})
    exp_df = add_month_col(df[df["type"]=="expense"].copy())
    monthly = exp_df.groupby("month")["amount"].sum().reset_index()
    monthly.columns = ["month","amount"]
    monthly = monthly.sort_values("month")
    actuals = monthly["amount"].tolist()
    avg = float(sum(actuals[-3:]) / min(3,len(actuals))) if actuals else 0
    future_months = []
    if len(monthly):
        last = monthly["month"].iloc[-1]
        for i in range(1,4):
            future_months.append(str(last + i))
    return jsonify({
        "months":   [str(m) for m in monthly["month"]] + future_months,
        "actual":   actuals + [None]*3,
        "forecast": [None]*len(actuals) + [round(avg*(1+0.02*i),2) for i in range(1,4)],
    })

@app.route("/api/recommendations")
@login_required
def api_recommendations():
    df = get_df()
    if df.empty:
        return jsonify({"tips":["Pehle kuch expenses add karo!"],"score":0})
    total_inc = float(df[df["type"]=="income"]["amount"].sum())
    total_exp = float(df[df["type"]=="expense"]["amount"].sum())
    savings_rate = ((total_inc-total_exp)/total_inc*100) if total_inc else 0
    exp_df = df[df["type"]=="expense"]
    top_cat = exp_df.groupby("category")["amount"].sum().idxmax() if not exp_df.empty else "N/A"
    tips = []
    if savings_rate < 20:
        tips.append(f"Savings rate {savings_rate:.1f}% hai — 20% target karo!")
    else:
        tips.append(f"Badiya! Savings rate {savings_rate:.1f}% hai.")
    tips.append(f"Sabse zyada spending '{top_cat}' mein hai — check karo.")
    if savings_rate >= 30: tips.append("Excellent savings! Investment ke baare mein socho.")
    score = min(100, int(savings_rate * 2))
    return jsonify({"tips": tips, "score": score})

@app.route("/download-excel")
@login_required
def download_excel():
    df = get_df()
    if df.empty:
        flash("Koi data nahi hai export karne ke liye!")
        return redirect(url_for("dashboard"))
    
    from flask import send_file
    import io
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Clean data
    df_export = df.copy()
    df_export['date'] = pd.to_datetime(df_export['date'], errors='coerce').dt.strftime('%Y-%m-%d')
    cols = ['date', 'category', 'type', 'amount', 'description']
    df_export = df_export[[c for c in cols if c in df_export.columns]]
    df_export.columns = ['Date', 'Category', 'Type', 'Amount (Rs.)', 'Description']

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='Transactions', startrow=2)
        wb = writer.book
        ws = writer.sheets['Transactions']

        # --- Title row ---
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = f'HomeFi — Transaction Report ({get_current_user()})'
        title_cell.font = Font(bold=True, size=14, color='FFFFFF')
        title_cell.fill = PatternFill('solid', fgColor='1a1a2e')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # --- Header row styling ---
        header_fill = PatternFill('solid', fgColor='4fc3f7')
        header_font = Font(bold=True, color='000000', size=11)
        for col in range(1, 6):
            cell = ws.cell(row=3, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # --- Data rows alternating colors ---
        light = PatternFill('solid', fgColor='1e1e3a')
        dark  = PatternFill('solid', fgColor='252545')
        green_font = Font(color='69f0ae', bold=True)
        red_font   = Font(color='ff5370', bold=True)

        for i, row in enumerate(ws.iter_rows(min_row=4, max_row=ws.max_row), start=0):
            for cell in row:
                cell.fill = light if i % 2 == 0 else dark
                cell.font = Font(color='e8eaf6')
                cell.alignment = Alignment(horizontal='center')
            # Amount column color
            type_cell   = ws.cell(row=row[0].row, column=3)
            amount_cell = ws.cell(row=row[0].row, column=4)
            if type_cell.value == 'income':
                amount_cell.font = green_font
            else:
                amount_cell.font = red_font

        # --- Column widths ---
        widths = [14, 16, 12, 16, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # --- Border ---
        thin = Side(style='thin', color='3a3a5c')
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    output.seek(0)
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'HomeFi_{get_current_user()}.xlsx')


@app.route("/download-pdf")
@login_required
def download_pdf():
    df = get_df()
    if df.empty:
        flash("Koi data nahi hai export karne ke liye!")
        return redirect(url_for("dashboard"))

    from flask import send_file
    import io
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                     Spacer, Image, PageBreak, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    GREEN = colors.HexColor('#2e7d32')
    RED = colors.HexColor('#c62828')
    BLUE = colors.HexColor('#1565c0')
    PURPLE = colors.HexColor('#6a1b9a')
    DARK = colors.HexColor('#1a1a2e')
    ACCENT = colors.HexColor('#4fc3f7')

    # ---------- Core numbers (all-time) ----------
    df_all = df.copy()
    total_income = float(df_all[df_all["type"]=="income"]["amount"].sum())
    total_expense = float(df_all[df_all["type"]=="expense"]["amount"].sum())
    net_savings = total_income - total_expense
    savings_rate = (net_savings/total_income*100) if total_income > 0 else 0

    df_m = add_month_col(df_all.copy())
    monthly = df_m.groupby(["month","type"])["amount"].sum().unstack(fill_value=0)
    months_sorted = sorted(monthly.index.astype(str))
    cur_month = months_sorted[-1]
    prev_month = months_sorted[-2] if len(months_sorted) > 1 else None

    def month_vals(m):
        if m is None:
            return 0.0, 0.0
        row = monthly.loc[monthly.index.astype(str) == m]
        inc = float(row["income"].values[0]) if "income" in row.columns and len(row) else 0.0
        exp = float(row["expense"].values[0]) if "expense" in row.columns and len(row) else 0.0
        return inc, exp

    cur_inc, cur_exp = month_vals(cur_month)
    prev_inc, prev_exp = month_vals(prev_month)
    cur_savings = cur_inc - cur_exp
    prev_savings = prev_inc - prev_exp
    exp_change = ((cur_exp - prev_exp)/prev_exp*100) if prev_exp > 0 else 0
    inc_change = ((cur_inc - prev_inc)/prev_inc*100) if prev_inc > 0 else 0
    sav_change = ((cur_savings - prev_savings)/abs(prev_savings)*100) if prev_savings != 0 else 0

    # ---------- Forecast (linear trend) ----------
    all_inc = [month_vals(m)[0] for m in months_sorted]
    all_exp = [month_vals(m)[1] for m in months_sorted]
    all_sav = [i-e for i,e in zip(all_inc, all_exp)]

    if len(all_sav) >= 2:
        x = np.arange(len(all_sav))
        slope, intercept = np.polyfit(x, all_sav, 1)
    else:
        slope, intercept = 0, all_sav[0] if all_sav else 0

    forecast_months = 3
    forecast_labels = []
    forecast_values = []
    last_period = pd.Period(months_sorted[-1])
    for i in range(1, forecast_months+1):
        fperiod = last_period + i
        forecast_labels.append(str(fperiod))
        fval = slope*(len(all_sav)-1+i) + intercept
        forecast_values.append(max(fval, 0))

    avg_monthly_savings = float(np.mean(all_sav)) if all_sav else 0

    # ---------- Goals ----------
    goals_res = supabase.table("goals").select("*").eq("user_id", get_current_user()).execute()
    all_goals_data = goals_res.data or []
    current_year_str = str(datetime.now().year)
    current_month_str = datetime.now().strftime("%Y-%m")
    yearly_goals = [g for g in all_goals_data if g["goal_type"]=="yearly" and g["period"]==current_year_str]
    monthly_goals = [g for g in all_goals_data if g["goal_type"]=="monthly" and g["period"]==current_month_str]

    df_year_filtered = df_m[df_m["month"].astype(str).str.startswith(current_year_str)]
    year_income = float(df_year_filtered[df_year_filtered["type"]=="income"]["amount"].sum())
    year_expense = float(df_year_filtered[df_year_filtered["type"]=="expense"]["amount"].sum())
    year_savings = year_income - year_expense

    # ---------- Category breakdown ----------
    cat_data = df_all[df_all["type"]=="expense"].groupby("category")["amount"].sum().sort_values(ascending=False)
    cur_cat = df_m[(df_m["type"]=="expense") & (df_m["month"].astype(str)==cur_month)].groupby("category")["amount"].sum().sort_values(ascending=False)
    prev_cat = df_m[(df_m["type"]=="expense") & (df_m["month"].astype(str)==prev_month)].groupby("category")["amount"].sum() if prev_month else pd.Series(dtype=float)

    # ---------- PDF setup ----------
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter,
                             topMargin=0.45*inch, bottomMargin=0.45*inch,
                             leftMargin=0.5*inch, rightMargin=0.5*inch)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('TitleStyle', parent=styles['Title'],
                                  textColor=DARK, fontSize=21, spaceAfter=2)
    sub_style = ParagraphStyle('SubStyle', parent=styles['Normal'],
                                textColor=colors.HexColor('#777777'), fontSize=9)
    heading_style = ParagraphStyle('HeadStyle', parent=styles['Heading2'],
                                    textColor=DARK, fontSize=13.5, spaceBefore=14, spaceAfter=6)
    subheading_style = ParagraphStyle('SubHeadStyle', parent=styles['Heading3'],
                                       textColor=colors.HexColor('#333355'), fontSize=11, spaceBefore=8, spaceAfter=4)
    body_style = ParagraphStyle('BodyStyle', parent=styles['Normal'], fontSize=9.5, leading=14)
    tip_style = ParagraphStyle('TipStyle', parent=styles['Normal'], fontSize=9.5, leading=15, leftIndent=4)

    elements = []

    # ---------------- HEADER ----------------
    elements.append(Paragraph("HomeFi — Complete Financial Analysis Report", title_style))
    elements.append(Paragraph(
        f"User: {get_current_user()} &nbsp;|&nbsp; Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} "
        f"&nbsp;|&nbsp; Period covered: {months_sorted[0]} to {months_sorted[-1]} "
        f"({len(months_sorted)} months)",
        sub_style))
    elements.append(Spacer(1, 10))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#dddddd')))
    elements.append(Spacer(1, 10))

    # ---------------- OVERALL SUMMARY ----------------
    elements.append(Paragraph("1. Overall Summary (All Months)", heading_style))
    summary_data = [
        ["Total Income", "Total Expense", "Net Savings (Fayda)", "Savings Rate"],
        [f"Rs.{total_income:,.0f}", f"Rs.{total_expense:,.0f}",
         f"Rs.{net_savings:,.0f}" if net_savings >= 0 else f"-Rs.{abs(net_savings):,.0f}",
         f"{savings_rate:.1f}%"]
    ]
    summary_table = Table(summary_data, colWidths=[1.7*inch]*4)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), DARK),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTSIZE', (0,1), (-1,1), 14),
        ('FONTNAME', (0,1), (-1,1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0,1), (0,1), GREEN),
        ('TEXTCOLOR', (1,1), (1,1), RED),
        ('TEXTCOLOR', (2,1), (2,1), GREEN if net_savings >= 0 else RED),
        ('TEXTCOLOR', (3,1), (3,1), PURPLE),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        f"<b>Fayda/Nuksaan status:</b> "
        + (f"Overall aapko <font color='#2e7d32'><b>Rs.{net_savings:,.0f} ka fayda</b></font> hua hai "
           f"in {len(months_sorted)} mahino mein."
           if net_savings >= 0 else
           f"Overall aapko <font color='#c62828'><b>Rs.{abs(net_savings):,.0f} ka nuksaan</b></font> hua hai "
           f"in {len(months_sorted)} mahino mein."),
        body_style))
    elements.append(Spacer(1, 10))

    # ---------------- MONTH-BY-MONTH ----------------
    elements.append(Paragraph("2. Month-by-Month Breakdown", heading_style))
    mb_data = [["Month", "Income", "Expense", "Savings (Fayda/Nuksaan)", "Status"]]
    for m in months_sorted:
        inc, exp = month_vals(m)
        sav = inc - exp
        status = "Fayda" if sav >= 0 else "Nuksaan"
        mb_data.append([m, f"Rs.{inc:,.0f}", f"Rs.{exp:,.0f}",
                         f"Rs.{sav:,.0f}" if sav >= 0 else f"-Rs.{abs(sav):,.0f}", status])

    mb_table = Table(mb_data, colWidths=[1.1*inch, 1.4*inch, 1.4*inch, 1.7*inch, 1*inch])
    mb_style = [
        ('BACKGROUND', (0,0), (-1,0), ACCENT),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (1,0), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('FONTSIZE', (0,0), (-1,-1), 9.5),
    ]
    for i, m in enumerate(months_sorted, start=1):
        inc, exp = month_vals(m)
        sav = inc - exp
        color = GREEN if sav >= 0 else RED
        mb_style.append(('TEXTCOLOR', (3,i), (4,i), color))
        mb_style.append(('FONTNAME', (3,i), (4,i), 'Helvetica-Bold'))
        bg = colors.HexColor('#f5f5f5') if i % 2 == 0 else colors.white
        mb_style.append(('BACKGROUND', (0,i), (2,i), bg))
    mb_table.setStyle(TableStyle(mb_style))
    elements.append(mb_table)
    elements.append(Spacer(1, 10))

    # ---------------- MONTHLY TREND CHART ----------------
    elements.append(Paragraph("3. Monthly Trend — Income vs Expense vs Savings", heading_style))
    fig, ax = plt.subplots(figsize=(6.6, 2.9))
    ax.plot(months_sorted, all_inc, marker='o', label='Income', color='#2e7d32', linewidth=2)
    ax.plot(months_sorted, all_exp, marker='o', label='Expense', color='#c62828', linewidth=2)
    ax.plot(months_sorted, all_sav, marker='o', label='Savings', color='#1565c0', linewidth=2, linestyle='--')
    ax.axhline(0, color='#999999', linewidth=0.8)
    ax.legend(fontsize=8)
    ax.tick_params(labelsize=8)
    ax.set_ylabel("Rs.", fontsize=8)
    for spine in ['top','right']:
        ax.spines[spine].set_visible(False)
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=130)
    plt.close(fig)
    buf.seek(0)
    elements.append(Image(buf, width=6.6*inch, height=2.9*inch))
    elements.append(Spacer(1, 10))

    # ---------------- CATEGORY BREAKDOWN ----------------
    if not cat_data.empty:
        elements.append(Paragraph("4. Category-wise Spending (Overall)", heading_style))
        fig, ax = plt.subplots(figsize=(6.6, 2.9))
        palette = ['#4fc3f7','#69f0ae','#ff5370','#ffd740','#b388ff','#ff8a65','#80cbc4','#f48fb1']
        ax.barh(cat_data.index[::-1], cat_data.values[::-1], color=palette[:len(cat_data)][::-1])
        ax.set_xlabel("Amount (Rs.)", fontsize=8)
        ax.tick_params(labelsize=8)
        for spine in ['top','right']:
            ax.spines[spine].set_visible(False)
        plt.tight_layout()
        buf2 = io.BytesIO()
        plt.savefig(buf2, format='png', dpi=130)
        plt.close(fig)
        buf2.seek(0)
        elements.append(Image(buf2, width=6.6*inch, height=2.9*inch))
        elements.append(Spacer(1, 4))

        cat_table_data = [["Category", "Total Spent", "% of Expenses"]]
        for cat, amt in cat_data.items():
            pct = amt/total_expense*100
            cat_table_data.append([cat, f"Rs.{amt:,.0f}", f"{pct:.1f}%"])
        cat_table = Table(cat_table_data, colWidths=[2.5*inch, 2*inch, 2*inch])
        cat_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), DARK),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN', (1,0), (-1,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
            ('FONTSIZE', (0,0), (-1,-1), 9.5),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]))
        elements.append(cat_table)

    elements.append(PageBreak())

    # ---------------- MONTH-OVER-MONTH COMPARISON ----------------
    if prev_month:
        elements.append(Paragraph("5. Month-over-Month Comparison", heading_style))
        elements.append(Paragraph(
            f"<b>{prev_month}</b> vs <b>{cur_month}</b> ka comparison.", body_style))
        elements.append(Spacer(1, 6))

        comp_data = [
            ["Metric", prev_month, cur_month, "Change", "Fayda/Nuksaan"],
            ["Income", f"Rs.{prev_inc:,.0f}", f"Rs.{cur_inc:,.0f}",
             f"{'+' if inc_change>=0 else ''}{inc_change:.1f}%",
             "Fayda" if inc_change >= 0 else "Nuksaan"],
            ["Expense", f"Rs.{prev_exp:,.0f}", f"Rs.{cur_exp:,.0f}",
             f"{'+' if exp_change>=0 else ''}{exp_change:.1f}%",
             "Nuksaan" if exp_change > 0 else "Fayda"],
            ["Savings", f"Rs.{prev_savings:,.0f}", f"Rs.{cur_savings:,.0f}",
             f"{'+' if sav_change>=0 else ''}{sav_change:.1f}%",
             "Fayda" if sav_change >= 0 else "Nuksaan"],
        ]
        comp_table = Table(comp_data, colWidths=[1.3*inch, 1.4*inch, 1.4*inch, 1.1*inch, 1.2*inch])
        comp_style = [
            ('BACKGROUND', (0,0), (-1,0), ACCENT),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN', (1,0), (-1,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('FONTSIZE', (0,0), (-1,-1), 9.5),
        ]
        for i, row in enumerate(comp_data[1:], start=1):
            color = GREEN if row[4] == "Fayda" else RED
            comp_style.append(('TEXTCOLOR', (3,i), (4,i), color))
            comp_style.append(('FONTNAME', (3,i), (4,i), 'Helvetica-Bold'))
        comp_table.setStyle(TableStyle(comp_style))
        elements.append(comp_table)
        elements.append(Spacer(1, 8))

        # Category-wise comparison
        elements.append(Paragraph("Category-wise change (vs previous month)", subheading_style))
        all_cats = sorted(set(cur_cat.index) | set(prev_cat.index))
        cc_data = [["Category", prev_month, cur_month, "Change"]]
        for c in all_cats:
            pv = float(prev_cat.get(c, 0))
            cv = float(cur_cat.get(c, 0))
            if pv > 0:
                chg = (cv-pv)/pv*100
                chg_str = f"{'+' if chg>=0 else ''}{chg:.0f}%"
            elif cv > 0:
                chg_str = "New"
            else:
                chg_str = "-"
            cc_data.append([c, f"Rs.{pv:,.0f}", f"Rs.{cv:,.0f}", chg_str])
        cc_table = Table(cc_data, colWidths=[2*inch, 1.6*inch, 1.6*inch, 1.2*inch])
        cc_style = [
            ('BACKGROUND', (0,0), (-1,0), DARK),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN', (1,0), (-1,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
            ('FONTSIZE', (0,0), (-1,-1), 9.5),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]
        for i, row in enumerate(cc_data[1:], start=1):
            if row[3].startswith('+'):
                cc_style.append(('TEXTCOLOR', (3,i), (3,i), RED))
            elif row[3].startswith('-') and row[3] != "-":
                cc_style.append(('TEXTCOLOR', (3,i), (3,i), GREEN))
        cc_table.setStyle(TableStyle(cc_style))
        elements.append(cc_table)
        elements.append(Spacer(1, 12))

    # ---------------- SAVINGS GOALS ----------------
    if yearly_goals or monthly_goals:
        elements.append(Paragraph("6. Savings Goals", heading_style))

        if yearly_goals:
            elements.append(Paragraph(f"Yearly Goals ({current_year_str})", subheading_style))
            goal_data = [["Goal", "Target", "Saved (Year)", "Remaining", "Progress"]]
            for g in yearly_goals:
                target = float(g["target_amount"])
                remaining = max(target - year_savings, 0)
                progress = min((year_savings/target*100), 100) if target > 0 else 0
                goal_data.append([
                    g["title"], f"Rs.{target:,.0f}",
                    f"Rs.{year_savings:,.0f}" if year_savings >= 0 else f"-Rs.{abs(year_savings):,.0f}",
                    f"Rs.{remaining:,.0f}", f"{progress:.1f}%"
                ])
            goal_table = Table(goal_data, colWidths=[1.6*inch, 1.1*inch, 1.2*inch, 1.2*inch, 0.9*inch])
            goal_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), PURPLE),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('ALIGN', (1,0), (-1,-1), 'CENTER'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
                ('FONTSIZE', (0,0), (-1,-1), 9),
                ('TOPPADDING', (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ]))
            elements.append(goal_table)
            elements.append(Spacer(1, 8))

        if monthly_goals:
            elements.append(Paragraph(f"Monthly Goals ({current_month_str})", subheading_style))
            goal_data = [["Goal", "Target", "Saved (Month)", "Remaining", "Progress"]]
            for g in monthly_goals:
                target = float(g["target_amount"])
                remaining = max(target - net_savings, 0)
                progress = min((net_savings/target*100), 100) if target > 0 else 0
                goal_data.append([
                    g["title"], f"Rs.{target:,.0f}",
                    f"Rs.{net_savings:,.0f}" if net_savings >= 0 else f"-Rs.{abs(net_savings):,.0f}",
                    f"Rs.{remaining:,.0f}", f"{progress:.1f}%"
                ])
            goal_table = Table(goal_data, colWidths=[1.6*inch, 1.1*inch, 1.2*inch, 1.2*inch, 0.9*inch])
            goal_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), ACCENT),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('ALIGN', (1,0), (-1,-1), 'CENTER'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
                ('FONTSIZE', (0,0), (-1,-1), 9),
                ('TOPPADDING', (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ]))
            elements.append(goal_table)
            elements.append(Spacer(1, 8))

    # ---------------- FORECAST ----------------
    elements.append(Paragraph("7. Forecast — Next 3 Months", heading_style))
    elements.append(Paragraph(
        "Pichle mahino ke trend ke aadhar par (simple linear projection), agle 3 mahino "
        "ki estimated savings:", body_style))
    elements.append(Spacer(1, 4))

    fig, ax = plt.subplots(figsize=(6.6, 2.9))
    ax.plot(months_sorted, all_sav, marker='o', label='Actual Savings', color='#1565c0', linewidth=2)
    combined_labels = months_sorted + forecast_labels
    combined_values = all_sav + forecast_values
    ax.plot(combined_labels[len(months_sorted)-1:], combined_values[len(months_sorted)-1:],
            marker='o', label='Forecast', color='#ff8a65', linewidth=2, linestyle='--')
    ax.axhline(0, color='#999999', linewidth=0.8)
    ax.legend(fontsize=8)
    ax.tick_params(labelsize=8, rotation=20)
    ax.set_ylabel("Savings (Rs.)", fontsize=8)
    for spine in ['top','right']:
        ax.spines[spine].set_visible(False)
    plt.tight_layout()
    buf4 = io.BytesIO()
    plt.savefig(buf4, format='png', dpi=130)
    plt.close(fig)
    buf4.seek(0)
    elements.append(Image(buf4, width=6.6*inch, height=2.9*inch))
    elements.append(Spacer(1, 4))

    fc_data = [["Month", "Projected Savings"]]
    for lbl, val in zip(forecast_labels, forecast_values):
        fc_data.append([lbl, f"Rs.{val:,.0f}"])
    fc_table = Table(fc_data, colWidths=[2.5*inch, 2.5*inch])
    fc_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), ACCENT),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (1,0), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
        ('FONTSIZE', (0,0), (-1,-1), 9.5),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TEXTCOLOR', (1,1), (1,-1), BLUE),
        ('FONTNAME', (1,1), (1,-1), 'Helvetica-Bold'),
    ]))
    elements.append(fc_table)
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(
        "<i>Note: Forecast ek simple trend-based estimate hai, actual results income/expenses "
        "mein real changes ke saath vary kar sakte hain.</i>", sub_style))
    elements.append(PageBreak())

    # ---------------- TIPS ----------------
    elements.append(Paragraph("8. Key Insights & Tips", heading_style))

    tips = []
    if not cat_data.empty:
        top_cat = cat_data.index[0]
        top_pct = cat_data.iloc[0]/total_expense*100
        tips.append(f"<b>Sabse zyada spending category:</b> '{top_cat}' — Rs.{cat_data.iloc[0]:,.0f} "
                    f"({top_pct:.1f}% of total expenses). Yahan budget set karne se savings improve "
                    f"ho sakti hai.")

    if prev_month:
        if exp_change > 0:
            tips.append(f"<b>Expense trend:</b> Pichle mahine ({prev_month}) ke comparison mein expenses "
                        f"<font color='#c62828'>{exp_change:.1f}% badhe</font> hain.")
        else:
            tips.append(f"<b>Expense trend:</b> Pichle mahine ({prev_month}) ke comparison mein expenses "
                        f"<font color='#2e7d32'>{abs(exp_change):.1f}% kam</font> hue hain — accha sign hai.")

    if savings_rate >= 20:
        tips.append(f"<b>Savings Rate:</b> {savings_rate:.1f}% — bahut healthy hai (recommended: 20%+).")
    elif savings_rate >= 0:
        tips.append(f"<b>Savings Rate:</b> {savings_rate:.1f}% — target rakho 20% ka, discretionary "
                    f"expenses mein 10-15% cut karne se ye possible hai.")
    else:
        tips.append(f"<b>Savings Rate:</b> {savings_rate:.1f}% — expenses income se zyada hain. "
                    f"Urgent action chahiye: non-essential categories mein cut karo.")

    if prev_month:
        all_cats2 = sorted(set(cur_cat.index) | set(prev_cat.index))
        increases = []
        for c in all_cats2:
            pv = float(prev_cat.get(c, 0))
            cv = float(cur_cat.get(c, 0))
            if pv > 0 and cv > pv:
                increases.append((c, cv-pv, (cv-pv)/pv*100))
        if increases:
            increases.sort(key=lambda x: x[1], reverse=True)
            big_c, big_amt, big_pct = increases[0]
            tips.append(f"<b>Biggest jump:</b> '{big_c}' category mein expense Rs.{big_amt:,.0f} "
                        f"({big_pct:.0f}%) badha hai pichle mahine se.")

    tips.append("<b>General tip:</b> Har mahine budget set karo har category ke liye, aur regularly "
                "transactions add karte raho — isse trends aur forecast zyada accurate honge.")

    for t in tips:
        elements.append(Paragraph(f"• {t}", tip_style))
        elements.append(Spacer(1, 3))
    elements.append(Spacer(1, 12))

    # ---------------- FULL TRANSACTION HISTORY ----------------
    elements.append(Paragraph("9. Full Transaction History", heading_style))
    df_sorted = df_all.sort_values('date', ascending=False)
    table_data = [["Date", "Month", "Category", "Type", "Amount (Rs.)", "Description"]]
    for _, row in df_sorted.iterrows():
        table_data.append([
            row['date'].strftime('%Y-%m-%d'), str(row['date'].to_period('M')), row['category'],
            row['type'].capitalize(), f"{row['amount']:,.0f}", str(row.get('description',''))[:30]
        ])

    trans_table = Table(table_data,
                         colWidths=[0.85*inch, 0.7*inch, 1.05*inch, 0.75*inch, 0.95*inch, 1.7*inch],
                         repeatRows=1)
    style_cmds = [
        ('BACKGROUND', (0,0), (-1,0), DARK),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8.5),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
        ('ALIGN', (4,0), (4,-1), 'RIGHT'),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]
    for i, row in enumerate(table_data[1:], start=1):
        bg = colors.HexColor('#f5f5f5') if i % 2 == 0 else colors.white
        style_cmds.append(('BACKGROUND', (0,i), (-1,i), bg))
        color = GREEN if row[3] == 'Income' else RED
        style_cmds.append(('TEXTCOLOR', (4,i), (4,i), color))
        style_cmds.append(('FONTNAME', (4,i), (4,i), 'Helvetica-Bold'))

    trans_table.setStyle(TableStyle(style_cmds))
    elements.append(trans_table)

    doc.build(elements)
    output.seek(0)
    return send_file(output, mimetype='application/pdf', as_attachment=True,
                     download_name=f'HomeFi_Report_{get_current_user()}.pdf')

@app.route("/delete-transaction/<int:transaction_id>")
@login_required
def delete_transaction(transaction_id):
    supabase.table("expenses").delete().eq("id", transaction_id).eq("user_id", get_current_user()).execute()
    flash("Transaction deleted!", "success")
    return redirect(url_for("dashboard"))

@app.route("/delete-account")
@login_required
def delete_account():
    user = get_current_user()
    supabase.table("expenses").delete().eq("user_id", user).execute()
    supabase.table("users").delete().eq("username", user).execute()
    session.clear()
    flash("Account deleted successfully!", "success")
    return redirect(url_for("login"))

@app.route("/clear-transactions")
@login_required
def clear_transactions():
    supabase.table("expenses").delete().eq("user_id", get_current_user()).execute()
    flash("Saare transactions delete ho gaye!", "success")
    return redirect(url_for("dashboard"))

@app.route("/history")
@login_required
def history():
    return render_template("history.html")

@app.route("/api/history")
@login_required
def api_history():
    df = get_df()
    if df.empty:
        return jsonify({"months": []})
    
    df = add_month_col(df)
    current_month = pd.Timestamp.now().to_period("M")
    
    months_data = []
    for month, group in df.groupby("month"):
        if month == current_month:
            continue  # skip current month
        income = float(group[group["type"]=="income"]["amount"].sum())
        expense = float(group[group["type"]=="expense"]["amount"].sum())
        months_data.append({
            "month": str(month),
            "income": income,
            "expense": expense,
            "savings": income - expense,
            "transactions": len(group)
        })
    
    months_data.sort(key=lambda x: x["month"], reverse=True)
    return jsonify({"months": months_data})

@app.route("/api/history/<month>")
@login_required
def api_history_month(month):
    df = get_df()
    if df.empty:
        return jsonify({"transactions": []})
    
    df = add_month_col(df)
    month_df = df[df["month"].astype(str) == month]
    
    transactions = []
    for _, row in month_df.iterrows():
        transactions.append({
            "date": row["date"].strftime("%Y-%m-%d"),
            "category": row["category"],
            "type": row["type"],
            "amount": float(row["amount"]),
            "description": row.get("description", "")
        })
    transactions.sort(key=lambda x: x["date"], reverse=True)
    return jsonify({"transactions": transactions})

@app.route("/goals")
@login_required
def goals_page():
    return render_template("goals.html")

@app.route("/api/goals")
@login_required
def api_goals():
    user = get_current_user()
    current_year = str(datetime.now().year)
    current_month = datetime.now().strftime("%Y-%m")

    res = supabase.table("goals").select("*").eq("user_id", user).execute()
    all_goals = res.data or []

    yearly = [g for g in all_goals if g["goal_type"]=="yearly" and g["period"]==current_year]
    monthly = [g for g in all_goals if g["goal_type"]=="monthly" and g["period"]==current_month]

    return jsonify({"yearly": yearly, "monthly": monthly})

@app.route("/api/goals/add", methods=["POST"])
@login_required
def api_goals_add():
    user = get_current_user()
    data = request.get_json()

    goal_type = data.get("goal_type")
    title = data.get("title")
    target_amount = float(data.get("target_amount"))

    if goal_type == "yearly":
        period = str(datetime.now().year)
    else:
        period = data.get("period") or datetime.now().strftime("%Y-%m")

    supabase.table("goals").insert({
        "user_id": user,
        "goal_type": goal_type,
        "period": period,
        "title": title,
        "target_amount": target_amount
    }).execute()

    return jsonify({"success": True})

@app.route("/api/goals/delete/<int:goal_id>", methods=["POST"])
@login_required
def api_goals_delete(goal_id):
    supabase.table("goals").delete().eq("id", goal_id).eq("user_id", get_current_user()).execute()
    return jsonify({"success": True})

@app.route("/budget")
@login_required
def budget_page():
    return render_template("budget.html")

@app.route("/api/budget-data")
@login_required
def api_budget_data():
    user = get_current_user()
    current_month = datetime.now().strftime("%Y-%m")

    df = get_df()
    df_m = add_month_col(df.copy())
    cur_spending = df_m[(df_m["type"]=="expense") & (df_m["month"].astype(str)==current_month)].groupby("category")["amount"].sum()

    res = supabase.table("budgets").select("*").eq("user_id", user).eq("month", current_month).execute()
    budgets = res.data or []
    budget_map = {b["category"]: float(b["amount"]) for b in budgets}

    all_cats = sorted(set(budget_map.keys()) | set(cur_spending.index))

    items = []
    for cat in all_cats:
        budget = budget_map.get(cat, 0)
        spent = float(cur_spending.get(cat, 0))
        items.append({
            "category": cat,
            "budget": budget,
            "spent": spent,
            "remaining": budget - spent,
            "pct": (spent/budget*100) if budget > 0 else 0
        })

    return jsonify({"month": current_month, "items": items})

@app.route("/api/budget/set", methods=["POST"])
@login_required
def api_budget_set():
    user = get_current_user()
    data = request.get_json()
    category = data.get("category", "").strip()
    amount = float(data.get("amount"))
    current_month = datetime.now().strftime("%Y-%m")

    if not category:
        return jsonify({"success": False, "error": "Category required"}), 400

    existing = supabase.table("budgets").select("*").eq("user_id", user).eq("category", category).eq("month", current_month).execute()
    if existing.data:
        supabase.table("budgets").update({"amount": amount}).eq("id", existing.data[0]["id"]).execute()
    else:
        supabase.table("budgets").insert({
            "user_id": user, "category": category, "amount": amount, "month": current_month
        }).execute()

    return jsonify({"success": True})
# ── Run ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)