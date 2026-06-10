"""
excel_exporter.py
-----------------
HomeFi — Export financial data and summaries to a formatted Excel workbook.
Creates 4 sheets: Raw Data, Monthly Summary, Category Summary, Dashboard.
"""

import os
import json
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint


REPORTS_DIR = "Reports"
os.makedirs(REPORTS_DIR, exist_ok=True)

# ── Colour palette ────────────────────────────────────────────────────────
C_DARK    = "1A1A2E"
C_ACCENT  = "00D4AA"
C_ACCENT2 = "F4A261"
C_GREEN   = "2DC653"
C_RED     = "E63946"
C_LGREY   = "F5F5F5"
C_MGREY   = "CCCCCC"
C_WHITE   = "FFFFFF"
C_BLUE    = "0D6EFD"

# ── Style helpers ─────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size,
                italic=italic, name="Arial")

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center")

def _right():
    return Alignment(horizontal="right", vertical="center")

def _border():
    thin = Side(style="thin", color=C_MGREY)
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _apply_header(ws, row, cols, text, span=None):
    """Write a section header cell."""
    cell = ws.cell(row=row, column=cols[0], value=text)
    cell.font      = _font(bold=True, color=C_WHITE, size=11)
    cell.fill      = _fill(C_DARK)
    cell.alignment = _center()
    if span and span > 1:
        ws.merge_cells(
            start_row=row, start_column=cols[0],
            end_row=row,   end_column=cols[0] + span - 1
        )

def _style_row(ws, row, col_start, col_end, fill_hex=None, bold=False,
               font_color="000000", align="left"):
    al = _center() if align == "center" else (_right() if align == "right" else _left())
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font      = _font(bold=bold, color=font_color)
        cell.alignment = al
        cell.border    = _border()
        if fill_hex:
            cell.fill = _fill(fill_hex)


# ── Sheet 1: Raw Data ─────────────────────────────────────────────────────
def _sheet_raw(wb, df):
    ws = wb.create_sheet("Raw Data")
    ws.sheet_view.showGridLines = False

    headers = ["Date", "Category", "Type", "Amount (Rs.)", "Description"]
    col_widths = [14, 16, 12, 16, 30]

    # Title
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value     = "HomeFi — All Transactions"
    t.font      = _font(bold=True, color=C_WHITE, size=13)
    t.fill      = _fill(C_DARK)
    t.alignment = _center()
    ws.row_dimensions[1].height = 28

    # Sub info
    ws.merge_cells("A2:E2")
    s = ws["A2"]
    s.value     = f"Exported: {datetime.now().strftime('%d %b %Y, %I:%M %p')}  |  Total Records: {len(df)}"
    s.font      = _font(color=C_MGREY, size=9, italic=True)
    s.fill      = _fill(C_DARK)
    s.alignment = _center()
    ws.row_dimensions[2].height = 16

    # Column headers
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font      = _font(bold=True, color=C_WHITE, size=10)
        cell.fill      = _fill("16213E")
        cell.alignment = _center()
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 20

    # Data rows
    for ri, (_, row) in enumerate(df.sort_values("date").iterrows(), 4):
        fill_hex = C_LGREY if ri % 2 == 0 else C_WHITE
        vals = [
            str(row["date"].date()),
            row["category"],
            row["type"],
            row["amount"],
            row.get("description", ""),
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = _fill(fill_hex)
            cell.border    = _border()
            cell.alignment = _left()
            if ci == 4:  # Amount column
                cell.number_format = '#,##0.00'
                cell.alignment     = _right()
                cell.font          = _font(
                    color=C_RED if row["type"] == "expense" else C_GREEN,
                    bold=True
                )
            else:
                cell.font = _font()

    ws.freeze_panes = "A4"


# ── Sheet 2: Monthly Summary ──────────────────────────────────────────────
def _sheet_monthly(wb, df):
    ws = wb.create_sheet("Monthly Summary")
    ws.sheet_view.showGridLines = False

    df = df.copy()
    df["month"] = df["date"].dt.to_period("M")
    monthly_inc = df[df["type"]=="income"].groupby("month")["amount"].sum()
    monthly_exp = df[df["type"]=="expense"].groupby("month")["amount"].sum()
    months      = sorted(set(monthly_inc.index) | set(monthly_exp.index))

    # Title
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value     = "HomeFi — Monthly Summary"
    t.font      = _font(bold=True, color=C_WHITE, size=13)
    t.fill      = _fill(C_DARK)
    t.alignment = _center()
    ws.row_dimensions[1].height = 28

    # Headers
    hdrs = ["Month", "Income (Rs.)", "Expenses (Rs.)", "Savings (Rs.)",
            "Savings Rate", "Status"]
    col_w = [14, 16, 16, 16, 14, 14]
    for ci, (h, w) in enumerate(zip(hdrs, col_w), 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font      = _font(bold=True, color=C_WHITE, size=10)
        cell.fill      = _fill("16213E")
        cell.alignment = _center()
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 20

    # Data
    for ri, m in enumerate(months, 3):
        inc  = monthly_inc.get(m, 0)
        exp  = monthly_exp.get(m, 0)
        sav  = inc - exp
        rate = (sav / inc * 100) if inc > 0 else 0
        status = ("Excellent" if rate >= 30 else
                  "Good"      if rate >= 20 else
                  "Average"   if rate >= 10 else "Low")
        status_color = (C_GREEN if rate >= 30 else
                        C_ACCENT if rate >= 20 else
                        C_ACCENT2 if rate >= 10 else C_RED)

        fill_hex = C_LGREY if ri % 2 == 0 else C_WHITE
        vals = [str(m), inc, exp, sav, rate / 100, status]

        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = _fill(fill_hex)
            cell.border    = _border()
            cell.alignment = _center()
            if ci == 1:
                cell.font = _font(bold=True)
            elif ci in (2, 3, 4):
                cell.number_format = '#,##0'
                cell.font = _font(
                    color=C_RED if (ci == 3 and exp > inc) else "000000"
                )
            elif ci == 5:
                cell.number_format = '0.0%'
                cell.font = _font(color=status_color, bold=True)
            elif ci == 6:
                cell.font = _font(bold=True, color=status_color)

    # Totals row
    tr = len(months) + 3
    ws.merge_cells(f"A{tr}:A{tr}")
    ws.cell(row=tr, column=1, value="TOTAL").font = _font(bold=True, color=C_WHITE)
    ws.cell(row=tr, column=1).fill = _fill(C_DARK)
    ws.cell(row=tr, column=1).alignment = _center()

    for ci in range(1, 7):
        cell = ws.cell(row=tr, column=ci)
        cell.fill   = _fill(C_DARK)
        cell.border = _border()
        cell.font   = _font(bold=True, color=C_WHITE)
        cell.alignment = _center()

    total_inc = sum(monthly_inc.values)
    total_exp = sum(monthly_exp.values)
    total_sav = total_inc - total_exp
    total_rate = (total_sav / total_inc * 100) if total_inc > 0 else 0

    ws.cell(row=tr, column=2, value=total_inc).number_format = '#,##0'
    ws.cell(row=tr, column=3, value=total_exp).number_format = '#,##0'
    ws.cell(row=tr, column=4, value=total_sav).number_format = '#,##0'
    ws.cell(row=tr, column=5, value=total_rate/100).number_format = '0.0%'
    ws.cell(row=tr, column=6, value="Overall")
    for ci in range(2, 7):
        ws.cell(row=tr, column=ci).fill      = _fill(C_DARK)
        ws.cell(row=tr, column=ci).font      = _font(bold=True, color=C_ACCENT)
        ws.cell(row=tr, column=ci).alignment = _center()
        ws.cell(row=tr, column=ci).border    = _border()

    ws.freeze_panes = "A3"

    # Bar chart
    chart = BarChart()
    chart.type    = "col"
    chart.title   = "Monthly Income vs Expenses"
    chart.y_axis.title = "Amount (Rs.)"
    chart.x_axis.title = "Month"
    chart.width   = 20
    chart.height  = 12

    inc_ref = Reference(ws, min_col=2, min_row=2, max_row=len(months)+2)
    exp_ref = Reference(ws, min_col=3, min_row=2, max_row=len(months)+2)
    chart.add_data(inc_ref, titles_from_data=True)
    chart.add_data(exp_ref, titles_from_data=True)
    chart.series[0].graphicalProperties.solidFill = C_ACCENT
    chart.series[1].graphicalProperties.solidFill = C_RED

    cats = Reference(ws, min_col=1, min_row=3, max_row=len(months)+2)
    chart.set_categories(cats)
    ws.add_chart(chart, f"A{tr + 3}")


# ── Sheet 3: Category Summary ─────────────────────────────────────────────
def _sheet_category(wb, df):
    ws = wb.create_sheet("Category Summary")
    ws.sheet_view.showGridLines = False

    exp_df = df[df["type"] == "expense"]
    cat_totals = exp_df.groupby("category")["amount"].sum().sort_values(ascending=False)
    total_exp  = cat_totals.sum()

    # Title
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value     = "HomeFi — Category-wise Expense Breakdown"
    t.font      = _font(bold=True, color=C_WHITE, size=13)
    t.fill      = _fill(C_DARK)
    t.alignment = _center()
    ws.row_dimensions[1].height = 28

    hdrs   = ["Category", "Total (Rs.)", "% of Expenses", "Avg/Month (Rs.)", "Transactions"]
    col_ws = [18, 16, 16, 16, 14]
    months = df["date"].dt.to_period("M").nunique()

    for ci, (h, w) in enumerate(zip(hdrs, col_ws), 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font      = _font(bold=True, color=C_WHITE, size=10)
        cell.fill      = _fill("16213E")
        cell.alignment = _center()
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w

    cat_colors = [
        C_ACCENT, C_ACCENT2, C_RED, "457B9D", C_GREEN,
        "9B5DE5", "F15BB5", "FEE440", "00BBF9", "8338EC"
    ]

    for ri, (cat, amt) in enumerate(cat_totals.items(), 3):
        pct    = amt / total_exp if total_exp > 0 else 0
        avg    = amt / months if months > 0 else amt
        txns   = len(exp_df[exp_df["category"] == cat])
        color  = cat_colors[(ri - 3) % len(cat_colors)]
        fill_h = C_LGREY if ri % 2 == 0 else C_WHITE

        vals = [cat, amt, pct, avg, txns]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill   = _fill(fill_h)
            cell.border = _border()
            if ci == 1:
                cell.font      = _font(bold=True, color=color)
                cell.alignment = _left()
            elif ci in (2, 4):
                cell.number_format = '#,##0'
                cell.font          = _font()
                cell.alignment     = _right()
            elif ci == 3:
                cell.number_format = '0.0%'
                cell.font          = _font(bold=True, color=color)
                cell.alignment     = _center()
            else:
                cell.font      = _font()
                cell.alignment = _center()

    ws.freeze_panes = "A3"

    # Pie chart
    pie = PieChart()
    pie.title  = "Expense Distribution by Category"
    pie.width  = 18
    pie.height = 14

    data = Reference(ws, min_col=2, min_row=2, max_row=len(cat_totals)+2)
    cats = Reference(ws, min_col=1, min_row=3, max_row=len(cat_totals)+2)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(cats)
    ws.add_chart(pie, "G2")


# ── Sheet 4: Dashboard ────────────────────────────────────────────────────
def _sheet_dashboard(wb, df):
    ws = wb.create_sheet("Dashboard", 0)   # First sheet
    ws.sheet_view.showGridLines = False

    total_inc  = df[df["type"]=="income"]["amount"].sum()
    total_exp  = df[df["type"]=="expense"]["amount"].sum()
    net_sav    = total_inc - total_exp
    sav_rate   = (net_sav / total_inc * 100) if total_inc > 0 else 0
    months     = df["date"].dt.to_period("M").nunique()
    records    = len(df)

    # Title banner
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value     = "HomeFi — Financial Dashboard"
    t.font      = _font(bold=True, color=C_WHITE, size=16)
    t.fill      = _fill(C_DARK)
    t.alignment = _center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:F2")
    s = ws["A2"]
    s.value     = f"Report Date: {datetime.now().strftime('%d %B %Y')}   |   Period: {months} months   |   Records: {records}"
    s.font      = _font(color=C_ACCENT, size=9, italic=True)
    s.fill      = _fill(C_DARK)
    s.alignment = _center()
    ws.row_dimensions[2].height = 18

    # KPI cards (row 4-7)
    kpis = [
        ("Total Income",   f"Rs. {total_inc:,.0f}",  C_GREEN,   "A"),
        ("Total Expenses", f"Rs. {total_exp:,.0f}",  C_RED,     "B"),
        ("Net Savings",    f"Rs. {net_sav:,.0f}",    C_ACCENT,  "C"),
        ("Savings Rate",   f"{sav_rate:.1f}%",       C_ACCENT2, "D"),
        ("Months Tracked", str(months),              C_BLUE,    "E"),
        ("Transactions",   str(records),             "9B5DE5",  "F"),
    ]

    for (label, value, color, col) in kpis:
        ws.column_dimensions[col].width = 20
        # Label row
        lc = ws[f"{col}4"]
        lc.value     = label
        lc.font      = _font(bold=True, color=C_WHITE, size=9)
        lc.fill      = _fill(color)
        lc.alignment = _center()
        lc.border    = _border()
        ws.row_dimensions[4].height = 20

        # Value row
        vc = ws[f"{col}5"]
        vc.value     = value
        vc.font      = _font(bold=True, color=color, size=14)
        vc.fill      = _fill(C_WHITE)
        vc.alignment = _center()
        vc.border    = _border()
        ws.row_dimensions[5].height = 30

        # Spacer
        sc = ws[f"{col}6"]
        sc.fill = _fill(C_LGREY)

    # Health badge
    ws.merge_cells("A7:F7")
    health_color = (C_GREEN  if sav_rate >= 30 else
                    C_ACCENT if sav_rate >= 20 else
                    C_ACCENT2 if sav_rate >= 10 else C_RED)
    health_txt   = ("EXCELLENT — Keep it up!" if sav_rate >= 30 else
                    "GOOD — On track."         if sav_rate >= 20 else
                    "AVERAGE — Can improve."   if sav_rate >= 10 else
                    "NEEDS IMPROVEMENT — Review your expenses.")
    hc = ws["A7"]
    hc.value     = f"Financial Health: {health_txt}"
    hc.font      = _font(bold=True, color=C_WHITE, size=11)
    hc.fill      = _fill(health_color)
    hc.alignment = _center()
    ws.row_dimensions[7].height = 24

    # Quick category table
    ws["A9"] = "TOP EXPENSE CATEGORIES"
    ws["A9"].font      = _font(bold=True, color=C_WHITE, size=10)
    ws["A9"].fill      = _fill(C_DARK)
    ws["A9"].alignment = _center()
    ws.merge_cells("A9:C9")

    exp_df     = df[df["type"]=="expense"]
    cat_top    = exp_df.groupby("category")["amount"].sum().sort_values(ascending=False).head(6)
    total_exp_ = cat_top.sum()

    ws.cell(row=10, column=1, value="Category").font = _font(bold=True, color=C_WHITE)
    ws.cell(row=10, column=2, value="Amount (Rs.)").font = _font(bold=True, color=C_WHITE)
    ws.cell(row=10, column=3, value="Share").font = _font(bold=True, color=C_WHITE)
    for c in range(1, 4):
        ws.cell(row=10, column=c).fill      = _fill("16213E")
        ws.cell(row=10, column=c).alignment = _center()
        ws.cell(row=10, column=c).border    = _border()

    for ri, (cat, amt) in enumerate(cat_top.items(), 11):
        pct      = amt / total_exp if total_exp > 0 else 0
        fill_hex = C_LGREY if ri % 2 == 0 else C_WHITE
        ws.cell(row=ri, column=1, value=cat).font      = _font(bold=True)
        ws.cell(row=ri, column=2, value=amt).number_format = '#,##0'
        ws.cell(row=ri, column=3, value=pct).number_format = '0.0%'
        for c in range(1, 4):
            ws.cell(row=ri, column=c).fill      = _fill(fill_hex)
            ws.cell(row=ri, column=c).alignment = _center()
            ws.cell(row=ri, column=c).border    = _border()

    ws.freeze_panes = "A3"


# ── Main export function ──────────────────────────────────────────────────
def export_to_excel(df, filename=None):
    """
    Export HomeFi data to a formatted Excel workbook.
    Returns the output file path.
    """
    if filename is None:
        ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"HomeFi_Export_{ts}.xlsx"

    out_path = os.path.join(REPORTS_DIR, filename)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    _sheet_dashboard(wb, df)
    _sheet_monthly(wb, df)
    _sheet_category(wb, df)
    _sheet_raw(wb, df)

    wb.save(out_path)
    print(f"\n  [OK] Excel file saved: {out_path}")
    print("  Sheets: Dashboard | Monthly Summary | Category Summary | Raw Data")
    return out_path


def run_excel_export(df):
    """Called from main.py."""
    print("\n" + "─" * 50)
    print("  EXPORT TO EXCEL")
    print("─" * 50)
    print("  Creating formatted Excel workbook...")
    path = export_to_excel(df)
    print(f"\n  Open '{path}' in Excel or LibreOffice.")


if __name__ == "__main__":
    from loader import load_data
    df = load_data("Data/expenses.csv")
    run_excel_export(df)